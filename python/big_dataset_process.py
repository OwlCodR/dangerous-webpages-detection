from googletrans import Translator
#from pymystem3 import Mystem
from openpyxl import load_workbook
import pymorphy2
import csv
import urllib
import string
import re
import codecs

# This file processing big .xlsx dataset
# Debug version

def get_sentences(text):
    print('get_sentences')
    text = text.replace('\n', ' ')
    text = text.replace('\t', ' ')
    text = text.split('.')
    text = [x for x in text if len(x) > 1]
    return text

def get_removed_punctuation(sentences):
    print('get_removed_punctuation')

    for i in range(len(sentences)):
        sentences[i] = sentences[i].lower()
        sentences[i] = ' '.join(re.findall(r'[А-я]+', sentences[i]))
    sentences = [x for x in sentences if len(x) > 3 or x == 'не' or x == 'ни']
    return sentences

def get_translated_sentences(sentences, translator):
    print('get_translated_sentences')

    for i in range(len(sentences)):
        sentences[i] = translator.translate(sentences[i], dest='ru').text
        print(sentences[i])

    return sentences

def get_lemmatize_sentences(sentences, morph):
    print('get_lemmatize_sentences')

    for i in range(len(sentences)):
        #lemmas = mystem.lemmatize(sentences[i])
        #sentences[i] = ''.join(lemmas).replace('\n', '')
        sentences[i] = ' '.join([morph.parse(word)[0].normal_form for word in sentences[i].split()])
        
        #print([sentences[i]])

    return sentences

def get_contained_key_words(key_words_path, output_delimiter, sentences, key_words):
    print('get_contained_key_words')

    removed = []

    for sentence in sentences:
        isKeyWord = False
        for key_word in key_words:
            if key_word in sentence:
                isKeyWord = True
                #print('isKeyWord -> ', key_word)
                break
        if not isKeyWord:
            sentences.remove(sentence)
            removed.append(sentence)
            print('Removed -> ', sentence)
    return [sentences, removed]
# один гадко одиночество тяжело устал некомфортно мочь кредит помогите достало 
# 450 слов

# Переведены только предложения, в которых есть слова
# из топ 450 слов из small_dataset.csv

'''
1. Стоит попробовать использовать все ключевые слова
2. Стоит самостоятельно добавить ключевые слова
3. Стоит проанализировать ключевые слова большого датасета
3.1 В датасете много мусорных слов
3.2 
4. Проверка по ключевым словам нужна, чтобы отбросить мусорные слова 
'''

if __name__ == "__main__":
    sentences_start_number = 2
    sentences_end_number = 64040
    # 32020 - half

    dataset_path = "python/server/dataset/big_dataset.xlsx"

    key_words_path = "python/server/dataset/key_words3.csv"
    key_words_delimiter = ';'

    removed_path = "python/server/dataset/big_dataset_removed.csv"
    removed_delimiter = ';'

    output_path = "python/server/dataset/big_dataset.csv"
    output_delimiter = ';'

    wb = load_workbook(dataset_path)
    sheet = wb.get_sheet_by_name('Sheet1')

    with open(output_path, mode="w", newline='', encoding='utf-8') as output:
        with open(key_words_path, encoding='utf-8') as key_words_csv:
            with open(removed_path, mode="w", newline='', encoding='utf-8') as removed_csv:
                file_reader = csv.reader(key_words_csv, delimiter = key_words_delimiter)
                file_writer = csv.writer(output, delimiter = output_delimiter)
                removed_writer = csv.writer(removed_csv, delimiter = removed_delimiter)
                #mystem = Mystem()
                morph = pymorphy2.MorphAnalyzer(lang='ru')

                translator = Translator()

                key_words = []
                for row in file_reader:
                    key_words.append(row[0])

                for i in range(sentences_start_number, sentences_end_number):
                    dataset_text = sheet.cell(row=i, column=1).value
                    dataset_markup = sheet.cell(row=i, column=2).value

                    #print(dataset_text)
                    sentences = get_sentences(str(dataset_text))
                    #print(sentences)
                    #sentences = get_translated_sentences(sentences, translator)
                    #print(sentences)
                    sentences = get_removed_punctuation(sentences)
                    #print(sentences)
                    sentences = get_lemmatize_sentences(sentences, morph)
                    #print(sentences)
                    if str(dataset_markup) == '1':
                        answer = get_contained_key_words(key_words_path, key_words_delimiter, sentences, key_words)
                        sentences = answer[0]
                        removed_sentences = answer[1]
                        if len(removed_sentences) > 0:
                            for removed_sentence in removed_sentences:
                                removed_writer.writerow([removed_sentence, dataset_markup])
                    #print(sentences)
                    for sentence in sentences:
                        file_writer.writerow([sentence, dataset_markup])
                    
                    print('Last -> ', i)