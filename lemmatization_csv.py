import csv
import codecs
from pymystem3 import Mystem
from openpyxl import load_workbook

# This file lemmatize text from .xlsx dataset to .csv file
# Using MyStem library by Yandex

def debug(dataset_text, lemma_text):
    print(dataset_text)
    print(' ==== TO ==== ')
    print(lemma_text)

def lemmatize_xlsx(lemma_text_count, dataset_path, output_path, output_delimiter):

    wb = load_workbook(dataset_path)
    sheet = wb.get_sheet_by_name('Sheet1')

    with codecs.open(output_path, mode="w", encoding='utf-8') as csv_file_lemma:
        file_writer = csv.writer(csv_file_lemma, delimiter = output_delimiter)
        mystem = Mystem()
        for i in range(4, lemma_text_count):
            dataset_text = sheet.cell(row=i, column=1).value
            dataset_markup = sheet.cell(row=i, column=2).value

            lemmas = mystem.lemmatize(dataset_text)
            lemma_text = ''.join(lemmas)

            debug(dataset_text, lemma_text)

            file_writer.writerow([lemma_text, dataset_markup])

if __name__ == "__main__":
    lemma_text_count = 7

    dataset_path = "dataset/big_dataset.xlsx"

    output_path = "dataset/lemma_big_dataset.csv"
    output_delimiter = '\t'

    lemmatize_xlsx(lemma_text_count, dataset_path, output_path, output_delimiter)